import io
import unittest
from unittest.mock import patch

import openpyxl

from app import app, _do_query
from tests.test_app_batch import FakeClient, VALID_ISBNS, collect_events


class ClassificationAPIIntegrationTest(unittest.TestCase):
    def book(self, code):
        return {"title": "测试书", "clc_code": code}

    def test_single_keeps_old_fields_and_adds_trace(self):
        with patch("app.query_isbn", return_value=self.book("I524.45")):
            response = app.test_client().post("/api/query", json={"isbn": VALID_ISBNS[0]})
        result = response.get_json()
        self.assertTrue(result["success"])
        for field in ["clc_code", "clc_name", "clc_path", "clc_path_str"]:
            self.assertIn(field, result)
        self.assertEqual(result["clc_status"], "supplemented")
        self.assertEqual(result["clc_parse"]["path"], result["clc_path"])

    def test_partial_path_does_not_fail_book_query_or_batch(self):
        with patch("app.query_isbn", return_value=self.book("TP311.12")) as query:
            events = collect_events([VALID_ISBNS[0], VALID_ISBNS[0]], FakeClient())
        result = events[-1]
        self.assertEqual(query.call_count, 1)
        self.assertEqual(result["success_count"], 2)
        self.assertEqual(result["results"][0]["clc_status"], "partial")
        self.assertNotIn("error_type", result["results"][0])

    def test_ancient_greece_is_consistent_in_single_batch_and_export(self):
        isbn = "9787020104420"
        with patch("app.query_isbn", return_value=self.book("I12(198.4)")):
            single = app.test_client().post("/api/query", json={"isbn": isbn}).get_json()
            batch = collect_events([isbn, isbn], FakeClient())[-1]
        self.assertEqual(batch["success_count"], 2)
        for result in [single] + batch["results"]:
            self.assertTrue(result["success"])
            self.assertEqual(result["clc_code"], "I12(198.4)")
            self.assertEqual(result["clc_path"], ["文学", "世界文学", "作品集", "诗歌集", "古代希腊"])
            self.assertEqual(result["clc_status"], "supplemented")
            self.assertEqual(result["clc_unparsed"], "")
            self.assertEqual(result["clc_code_status"], "historical")
            self.assertTrue(any("未鉴定" in warning for warning in result["clc_warnings"]))
        response = app.test_client().post("/api/export", json={"results": [single]})
        self.assertEqual(response.status_code, 200)
        workbook = openpyxl.load_workbook(io.BytesIO(response.data), data_only=False)
        sheet = workbook.active
        headers = {cell.value: cell.column for cell in sheet[1]}
        self.assertEqual(sheet.cell(2, headers["中图分类号"]).value, "I12(198.4)")
        self.assertEqual(sheet.cell(2, headers["分类路径"]).value, single["clc_path_str"])
        self.assertEqual(sheet.cell(2, headers["解析状态"]).value, "规则补全")
        self.assertEqual(sheet.cell(2, headers["号码状态"]).value, "含历史地区号")
        self.assertIn("未鉴定", sheet.cell(2, headers["解析提示"]).value)

    def test_parser_read_failure_does_not_retry_nlc(self):
        with patch("app.query_isbn", return_value=self.book("I524.45")) as query:
            with patch("clc_rules.engine.get_base", side_effect=OSError("模拟读取失败")):
                with self.assertLogs("clc_rules.engine", level="ERROR"):
                    result = _do_query(VALID_ISBNS[0])
        self.assertTrue(result["success"])
        self.assertEqual(result["clc_status"], "unavailable")
        self.assertEqual(query.call_count, 1)

    def test_export_uses_same_status_and_literal_unparsed_equals(self):
        with patch("app.query_isbn", return_value=self.book("I524.45=999")):
            result = _do_query(VALID_ISBNS[0])
        result["title"] = "=HYPERLINK(\"https://example.invalid\",\"not a formula\")"
        response = app.test_client().post("/api/export", json={"results": [result]})
        self.assertEqual(response.status_code, 200)
        workbook = openpyxl.load_workbook(io.BytesIO(response.data), data_only=False)
        sheet = workbook.active
        headers = {cell.value: cell.column for cell in sheet[1]}
        self.assertNotIn("完整分类路径", headers)
        for field, value in [("分类路径", result["clc_path_str"]), ("解析状态", "部分解析"),
                             ("未解析部分", "=999"), ("中图分类号", "I524.45=999")]:
            self.assertEqual(sheet.cell(2, headers[field]).value, value)
        for field in ["未解析部分", "书名"]:
            self.assertEqual(sheet.cell(2, headers[field]).data_type, "s")
        self.assertEqual(sheet.cell(2, headers["查询状态"]).value, "成功")

    def test_ten_headerless_excel_rows_are_preserved(self):
        workbook = openpyxl.Workbook()
        for index in range(10):
            workbook.active.append([VALID_ISBNS[index % len(VALID_ISBNS)]])
        content = io.BytesIO()
        workbook.save(content)
        content.seek(0)
        with patch("app.query_isbn", return_value=self.book("I524.45")):
            response = app.test_client().post("/api/batch", data={"file": (content, "测试.xlsx")})
        import json
        done = json.loads(response.data.decode().splitlines()[-1])
        self.assertEqual(done["total"], 10)
        self.assertEqual(len(done["results"]), 10)


if __name__ == "__main__":
    unittest.main()
